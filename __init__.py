import logging
import datetime
import codecs
import os
import sqlite3
import numpy as np
from sqlite3 import OperationalError
from copy import copy
from threading import Thread
import openpyxl
import pandas as pd
import xgboost as xgb
from threading import Thread
from flask import Flask, render_template, request, redirect, url_for, session
from flask import send_file, jsonify, after_this_request, make_response
from flask_bootstrap import Bootstrap
from flask_login import LoginManager, UserMixin, login_user, login_required
from flask_login import logout_user, current_user
from flask_mail import Mail, Message
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from werkzeug.security import generate_password_hash, check_password_hash
from wtforms import StringField, PasswordField, BooleanField, IntegerField, SelectField
from wtforms.validators import InputRequired, Length, Email, EqualTo
from wtforms.validators import ValidationError, NumberRange, DataRequired

app = Flask(__name__)

app.jinja_env.filters['zip'] = zip
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'
app.config['SQLALCHEMY_DATABASE_URI'] = \
    'sqlite:///diacompanion.db'
app.config['TESTING'] = False
app.config['DEBUG'] = True
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = ''
app.config['MAIL_PASSWORD'] = ''
app.config['MAIL_DEFAULT_SENDER'] = ('Еженедельник', '')
app.config['MAIL_MAX_EMAILS'] = None
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAIL_ASCII_ATTACHMENTS'] = False
app.config['SESSION_COOKIE_SAMESITE'] = "Lax"


Bootstrap(app)
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
mail = Mail(app)


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(15), unique=True)
    username1 = db.Column(db.String(80))
    email = db.Column(db.String(80), unique=True)
    password = db.Column(db.String(15))
    BMI = db.Column(db.String(80))
    doc = db.Column(db.String(80))


class LoginForm(FlaskForm):
    username = StringField('Логин или email',
                           validators=[InputRequired(message='Необходимо \
                                                     заполнить это поле'),
                                       Length(min=5,
                                              max=80,
                                              message='Необходимо минимум 5 \
                                                       символов')])
    password = PasswordField('Пароль', validators=[InputRequired(message='\
                                                                Необходимо \
                                                                заполнить \
                                                                это \
                                                                поле'),
                                                   Length(min=5,
                                                          max=15,
                                                          message='Пароль \
                                                                   должен \
                                                                   быть от 5 \
                                                                   до 15 \
                                                                   символов')])
    remember = BooleanField('Запомнить меня')


class RegisterForm(FlaskForm):
    email = StringField('Email',
                        validators=[InputRequired(message='Необходимо \
                                                           заполнить \
                                                           это поле'),
                                    Email(message='Неправильно введен \
                                                   email'),
                                    Length(max=80)])
    username = StringField('Логин',
                           validators=[InputRequired(message='Необходимо \
                                                              заполнить \
                                                              это поле'),
                                       Length(min=5,
                                              max=15,
                                              message='Никнейм \
                                                       должен \
                                                       быть от 5 \
                                                       до 15 \
                                                       символов')])
    username1 = StringField('ФИО пользователя',
                            validators=[InputRequired(message='Необходимо \
                                                               заполнить \
                                                               это поле'),
                                        Length(min=5,
                                               max=80,
                                               message='Необходимо минимум 5 \
                                                        символов')])
    password = PasswordField('Пароль',
                             validators=[InputRequired(message='Создайте \
                                                                ваш \
                                                                пароль'),
                                         Length(min=5, max=15,
                                                message='Пароль \
                                                         должен \
                                                         быть от \
                                                         5 до 15 \
                                                         символов')])
    password1 = PasswordField('Подтвердите пароль',
                              validators=[InputRequired(message='Необходимо \
                                                                 заполнить \
                                                                 это поле'),
                                          EqualTo(fieldname='password',
                                                  message='Пароли \
                                                           не совпадают')])
    weight = IntegerField('Вес, в кг',
                          validators=[InputRequired(message='Необходимо \
                                                            заполнить это \
                                                            поле'),
                                      NumberRange(min=0,
                                                  max=200,
                                                  message='Укажите свой \
                                                           реальный вес'),
                                      DataRequired(message='Введите целое \
                                                            число')])
    height = IntegerField('Рост, в см',
                          validators=[InputRequired(message='Необходимо \
                                                            заполнить это \
                                                            поле'),
                                      NumberRange(min=0,
                                                  max=250,
                                                  message='Укажите свой \
                                                           реальный рост'),
                                      DataRequired(message='Введите целое \
                                                            число')])
    select = SelectField(u'Лечащий врач', choices=[('pvpopova@ya.ru',
                                                    'Попова П.В.'),
                                                   ('aleksandra.tkachuk.1988@mail.comm',
                                                    'Ткачук А.С.'),
                                                   ('yanabolotko@gmail.com',
                                                    'Болотько Я.А.'),
                                                   ('aleksandra-dronova@yandex.ru',
                                                    'Дронова А.В.'),
                                                   ('elenavasukova2@gmail.com',
                                                    'Васюкова Е.А.'),
                                                   ('anopova.ann@gmail.com',
                                                    'Анопова А.Д.'),
                                                   ('andreigerasimov2704@gmail.com',
                                                    'Герасимов А.С.'),
                                                   ('tatarinova.maria@mail.ru',
                                                    'Татаринова М.В.'),
                                                   ('anna.datsiuk@mail.ru',
                                                    'Дацюк А.М.')])

    def validate_username(self, username):
        user = User.query.filter_by(username=username.data).first()
        if user is not None:
            raise ValidationError('Имя пользователя занято')

    def validate_email(self, email):
        email = User.query.filter_by(email=email.data).first()
        if email is not None:
            raise ValidationError('Email уже использовался')


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route('/')
def zero():
    # Перенаправляем на страницу входа/регистрации
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():   
    # Авторизация пользователя
    form = LoginForm()

    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None:
            user = User.query.filter_by(email=form.username.data).first()
        if user:
            if check_password_hash(user.password, form.password.data):
                login_user(user, remember=form.remember.data)
                return redirect(url_for('lk'))
        form.username.errors.append('')
        form.password.errors.append('Неверно введено имя пользователя \
                                    или пароль')
        form.password.data = ''
    return render_template('LO.html', form=form)


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    # Регистрация пользователя
    form = RegisterForm()

    if form.validate_on_submit():
        hashed_password = generate_password_hash(form.password.data,
                                                 method='sha256')
        BMI = form.weight.data/((form.height.data/100)*(form.height.data/100))
        BMIdata = str(BMI)
        new_user = User(username=form.username.data,
                        username1=form.username1.data, email=form.email.data,
                        password=hashed_password, BMI=BMIdata, doc=form.select.data)
        db.session.add(new_user)
        db.session.commit()
        db.session.close()
        return redirect(url_for('login'))
    return render_template('SU.html', form=form)


@app.route('/logout')
@login_required
def logout():
    # Выход из сети
    logout_user()
    return redirect(url_for('login'))


@app.route('/news')
@login_required
def news():
    # Главная страница
    path = os.path.dirname(os.path.abspath(__file__))
    db_2 = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db_2)
    cur = con.cursor()
    cur.execute("""SELECT food,libra FROM basket WHERE user_id = ?""",
                (session['user_id'],))
    result = cur.fetchall()
    len2 = len(result)+1
    con.close()

    return render_template("searching.html", result=result, len = len2)

@app.route('/onlinepredict', methods=['GET', 'POST'])
@login_required
def pred():
    if request.method == 'POST':
        jsoninfo = request.get_json()
        path = os.path.dirname(os.path.abspath(__file__))
        db_21 = os.path.join(path, 'diacompanion.db')
        model = os.path.join(path, 'model.model')
        con = sqlite3.connect(db_21)
        cur = con.cursor()
        cur.execute('''SELECT BMI FROM user WHERE id = ?''',(session['user_id'],))
        BMI0 = cur.fetchall()
        BMI0 = BMI0[0][0]
        nutr = list()
        for i in range(len(jsoninfo['foodname'])):
            cur.execute("""SELECT gi,carbo,prot,kr
                        FROM food WHERE name = ?""",
                        (jsoninfo['foodname'][i].split('//')[0],))
            nutrients = cur.fetchall()
            nutr.append(nutrients[0])
        cur.execute("""SELECT date,time,type,BG0,gi,carbo,prot,kr FROM favourites
                    WHERE user_id = ?""", (session["user_id"],))
        tb1 = cur.fetchall()
        con.close()

        tb1 = pd.DataFrame(tb1, columns=['date', 'time', 'types_food_n', 'BG0',
                                       'GI', 'carbo', 'prot', 'kr'])
        tb1['GI'] = pd.to_numeric(tb1['GI'], downcast='float')
        tb1['carbo'] = pd.to_numeric(tb1['carbo'], downcast='float')
        tb1['prot'] = pd.to_numeric(tb1['prot'], downcast='float')
        tb1['kr'] = pd.to_numeric(tb1['kr'], downcast='float')
        tb1['BG0'] = pd.to_numeric(tb1['BG0'], downcast='float')

        tb = pd.DataFrame(nutr, columns=['GI','carbo','prot','kr'])
        datenumb = jsoninfo['Date'].split('-')
        ddmmyy = '.'.join([datenumb[2],datenumb[1],datenumb[0]])
        date = pd.Series([ddmmyy]*len(tb['GI']))
        tb['date'] = date
        time = pd.Series([jsoninfo['Time']]*len(tb['GI']))
        tb['time'] = time
        typ_e = pd.Series([jsoninfo['Type']]*len(tb['GI']))
        tb['types_food_n'] = typ_e
        BG0 = pd.Series([jsoninfo['BG0']]*len(tb['GI']))
        tb['BG0'] = BG0
        tb = tb[['date','time','types_food_n','BG0','GI','carbo','prot','kr']]

        tb['GI'] = pd.to_numeric(tb['GI'], downcast='float')
        tb['carbo'] = pd.to_numeric(tb['carbo'], downcast='float')
        tb['prot'] = pd.to_numeric(tb['prot'], downcast='float')
        tb['kr'] = pd.to_numeric(tb['kr'], downcast='float')
        tb['BG0'] = pd.to_numeric(tb['BG0'], downcast='float')

        tb = pd.merge(left=tb, right=tb1, on=['date','time','types_food_n','BG0','GI','carbo','prot','kr'], how='outer')

        tb = tb.groupby(['date', 'time', 'types_food_n', 'BG0'],
                        as_index=False).sum()
        tb['GL'] = tb['GI']*tb['carbo']/100
        tb['DateTime'] = tb['date'] + ' ' + tb['time']
        tb['DateTime'] = pd.to_datetime(tb['DateTime'], format='%d.%m.%Y %H:%M')
        tb = tb.drop(['date', 'time', 'GI'], axis=1)
        prot = list()

        for i in range(len(tb['DateTime'])):
            start_date = tb['DateTime'][i]
            mask = (tb['DateTime']
                    <= start_date) & (tb['DateTime']
                                      >= (start_date
                                          - pd.Timedelta(value=6, unit='h')))
            prot_b6h = tb.loc[mask]['prot'].aggregate(np.sum)
            prot.append(prot_b6h)
        tb.insert(7, 'prot_b6h', prot, True)
        tb = tb.drop(['prot'], axis=1)

        BMI = list()
        for i in range(len(tb['DateTime'])):
            BMI.append(BMI0)
            if tb['types_food_n'][i] == 'Завтрак':
                tb['types_food_n'][i] = 1
            elif tb['types_food_n'][i] == 'Обед':
                tb['types_food_n'][i] = 2
            elif tb['types_food_n'][i] == 'Ужин':
                tb['types_food_n'][i] = 3
            else:
                tb['types_food_n'][i] = 4
        tb.insert(7, 'BMI', BMI, True)
        tb = tb.reindex(columns=["DateTime", "BG0", "GL", "carbo", "prot_b6h",
                                 "types_food_n", "kr", "BMI"])
        predict = list()
        for i in range(len(tb['DateTime'])):
            best_model = xgb.Booster()
            best_model.load_model(model)
            core_features = ["BG0", "gl", "carbo",
                             "prot_b6h", "types_food_n", "kr", "BMI"]
            X_test = [tb.iloc[i, 1:7].values.tolist()]
            predicted = best_model.predict(xgb.DMatrix(np.array(X_test)))
            predict.append(predicted[0])
        tb.insert(3, 'Предсказанный сахар после', predict, True)
        date3 = list()
        time3 = list()
        tb['Прием пищи'] = tb['types_food_n']

        for i in range(len(tb['DateTime'])):
            date3.append(tb['DateTime'][i].strftime('%d.%m.%Y'))
            time3.append(tb['DateTime'][i].strftime('%H:%M'))
            if tb['Прием пищи'][i] == 1:
                tb['Прием пищи'][i] = 'Завтрак'
            elif tb['Прием пищи'][i] == 2:
                tb['Прием пищи'][i] = 'Обед'
            elif tb['Прием пищи'][i] == 3:
                tb['Прием пищи'][i] = 'Ужин'
            else:
                tb['Прием пищи'][i] = 'Перекус'
        tb.insert(0, 'Дата', date3, True)
        tb.insert(1, 'Время', time3, True)
        tb = tb.drop(['DateTime'], axis=1)
        tb = tb.drop(['types_food_n'], axis=1)

        tb['Сахар до'] = tb['BG0']
        tb = tb.drop(['BG0'], axis=1)
        tb['Гликемическая нагрузка'] = tb['GL']
        tb = tb.drop(['GL'], axis=1)
        tb = tb.drop(['carbo'], axis=1)
        tb = tb.drop(['prot_b6h'], axis=1)
        tb = tb.drop(['kr'], axis=1)
        tb = tb.drop(["BMI"], axis=1)
        tb = tb[['Дата', 'Время', 'Прием пищи', 'Сахар до',
                 'Предсказанный сахар после', 'Гликемическая нагрузка']]

        mask1 = (tb['Дата'] == ddmmyy) & (tb['Время'] == jsoninfo['Time'])
        BG1 = tb.loc[mask1]['Предсказанный сахар после'].aggregate(np.sum)
        if (BG1 > 0.1)&(BG1 < 7):
            messag_e = 'УСК после еды в норме'
        elif BG1 < 0.1:
            messag_e = ''
        else:
            messag_e = 'УСК после еды превысит норму'
        list2 = jsonify({"BG1": messag_e})
        response = make_response(list2, 200)
    return response

@app.route('/search_page')
@login_required
def search_page():
    # Поисковая страница
    return render_template("searching.html")


@app.route('/searchlink/<string:search_string>')
@login_required
def searchlink(search_string):
    # Работа селекторного меню "выбрать категорию"
    path = os.path.dirname(os.path.abspath(__file__))
    db_3 = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db_3)
    cur = con.cursor()
    cur.execute("""SELECT name,_id
                FROM food WHERE category LIKE ?""", ('%{}%'.format(search_string),))
    result = cur.fetchall()
    list1 = pd.DataFrame(result, columns=['name','id'])
    list2 = list()
    for i in range(len(list1['name'])):
        cur.execute("""SELECT receipt,name
                    FROM recipes WHERE name LIKE ?""", (list1['name'][i],))
        receipt = cur.fetchall()
        try:
            list2.append(receipt[0])
        except IndexError:
            list2.append(('',''))
    con.close()
    list2 = pd.DataFrame(list2, columns=['receipt','name'])
    lis_t = pd.merge(left=list1, right=list2, on='name', how='left')
    lis_t = lis_t.replace(np.nan, '', regex=True)
    len1 = len(lis_t['receipt'])
    return render_template('searching_add.html', name=lis_t['name'], receipt=lis_t['receipt'],
                            id=lis_t['id'], len=len1)


@app.route('/search', methods=['GET', 'POST'])
@login_required
def search():
    # Основная функция сайта - поиск по базе данных
    if request.method == 'POST':
        search_string = request.form['input_query']
        search_string = search_string.capitalize()
        path = os.path.dirname(os.path.abspath(__file__))
        db_4 = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db_4)
        cur = con.cursor()

        cur.execute(""" SELECT category FROM foodGroups""")
        category_a = cur.fetchall()
        if (request.form['input_query'],) in category_a:
            cur.execute('''SELECT name,_id FROM food
                        WHERE category LIKE ?''', ('%{}%'.format(search_string),))
            result = cur.fetchall()
            list1 = pd.DataFrame(result, columns=['name','id'])
            list2 = list()
            for i in range(len(list1['name'])):
                cur.execute("""SELECT receipt,name
                            FROM recipes WHERE name LIKE ?""", (list1['name'][i],))
                receipt = cur.fetchall()
                try:
                    list2.append(receipt[0])
                except IndexError:
                    list2.append(('',''))
            list2 = pd.DataFrame(list2, columns=['receipt','name'])
            lis_t = pd.merge(left=list1, right=list2, on='name', how='left')
            lis_t = lis_t.replace(np.nan, '', regex=True)
            len1 = len(lis_t['receipt'])
        else:
            cur.execute('''SELECT name,_id FROM food
                        WHERE name LIKE ?
                        GROUP BY name''', ('%{}%'.format(search_string),))
            result = cur.fetchall()
            list1 = pd.DataFrame(result, columns=['name','id'])
            list2 = list()
            for i in range(len(list1['name'])):
                cur.execute("""SELECT receipt,name
                            FROM recipes WHERE name LIKE ?""", (list1['name'][i],))
                receipt = cur.fetchall()
                try:
                    list2.append(receipt[0])
                except IndexError:
                    list2.append(('',''))
            list2 = pd.DataFrame(list2, columns=['receipt','name'])
            lis_t = pd.merge(left=list1, right=list2, on='name', how='left')
            lis_t = lis_t.replace(np.nan, '', regex=True)
            len1 = len(lis_t['receipt'])
        con.close()
    return render_template('searching_add.html', name=lis_t['name'], receipt=lis_t['receipt'],
                           id=lis_t['id'], len=len1)


@app.route('/favourites', methods=['POST', 'GET'])
@login_required
def favour():
    # Добавляем блюда в предварительный список
    if request.method == 'POST':

        L1 = request.form.getlist('row')
        libra = request.form['libra']

        path = os.path.dirname(os.path.abspath(__file__))
        db_5 = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db_5)
        cur = con.cursor()

        for i in range(len(L1)):
            cur.execute(f"""INSERT INTO basket (user_id, food, libra)
                            VALUES (?,?,?)""",
                        (session['user_id'], L1[i], libra))
            con.commit()
        con.close()

    return redirect(url_for('news'))


@app.route('/favourites_dell', methods=['POST', 'GET'])
@login_required
def favour_dell():
    # Удаление ошибочных записей из "предварительного" списка
    if request.method == 'POST':
        flist = request.form.getlist('row')
        food = []
        libra = []
        for i in range(len(flist)):
            flist[i] = flist[i].split('//')
            food.append(flist[i][0])
            libra.append(flist[i][1])

        for i in range(len(food)):
            path = os.path.dirname(os.path.abspath(__file__))
            db_6 = os.path.join(path, 'diacompanion.db')
            con = sqlite3.connect(db_6)
            cur = con.cursor()
            cur.execute("""DELETE FROM basket WHERE user_id = ? AND food = ?
                        AND libra = ?""",
                        (session['user_id'], food[i], libra[i]))
            con.commit()
            con.close()
    return redirect(url_for('news'))


@app.route('/favourites_add', methods=['POST', 'GET'])
@login_required
def favour_add():
    # Добавляем блюда в основную базу данных и стираем временный список basket
    if request.method == 'POST':

        brf1 = datetime.time(7, 0)
        brf2 = datetime.time(11, 30)
        obed1 = datetime.time(12, 0)
        obed2 = datetime.time(15, 0)
        ujin1 = datetime.time(18, 0)
        ujin2 = datetime.time(22, 0)
        now = datetime.datetime.now().time()

        time = request.form['timer']
        if time == "":
            x = datetime.datetime.now().time()
            time = x.strftime("%R")
        else:
            x = datetime.datetime.strptime(time, "%H:%M")
            time = x.strftime("%R")

        date = request.form['calendar']
        if date == "":
            y = datetime.datetime.today().date()
            date = y.strftime("%d.%m.%Y")
            week_day = y.strftime("%A")
        else:
            y = datetime.datetime.strptime(date, "%Y-%m-%d")
            y = y.date()
            date = y.strftime("%d.%m.%Y")
            week_day = y.strftime("%A")

        if week_day == 'Monday':
            week_day = 'Понедельник'
        elif week_day == 'Tuesday':
            week_day = 'Вторник'
        elif week_day == 'Wednesday':
            week_day = 'Среда'
        elif week_day == 'Thursday':
            week_day = 'Четверг'
        elif week_day == 'Friday':
            week_day = 'Пятница'
        elif week_day == 'Saturday':
            week_day = 'Суббота'
        else:
            week_day = 'Воскресенье'
        typ = request.form['food_type']
        if typ == "Авто":
            if brf2 < now < brf1:
                typ = "Завтрак"
            elif obed1 < now < obed2:
                typ = "Обед"
            elif ujin1 < now < ujin2:
                typ = "Ужин"
            else:
                typ = "Перекус"

        path = os.path.dirname(os.path.abspath(__file__))
        db_7 = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db_7)
        cur = con.cursor()

        # Достаем названия и граммы из временно созданной корзины basket
        cur.execute("""SELECT food FROM basket WHERE user_id = ?""",
                    (session['user_id'],))
        L1 = cur.fetchall()

        cur.execute("""SELECT libra FROM basket WHERE user_id = ?""",
                    (session['user_id'],))
        libra = cur.fetchall()

        BG0 = request.form['sug']

        # Достаем все необходимые для диеты параметры
        elem1 = ['prot', 'carbo', 'fat', 'ec', 'water', 'mds', 'kr',
                 'pv', 'ok', 'zola', 'na', 'k', 'ca', 'mg', 'p',
                 'fe', 'a', 'kar', 're', 'b1', 'b2', 'rr', 'c', 'hol',
                 'nzhk', 'ne', 'te', 'gi']
        elem2 = ['prot', 'carbo', 'fat', 'energy', 'water', 'mds', 'kr',
                 'pv', 'ok', 'zola', 'na', 'k', 'ca', 'mg', 'p', 'fe', 'a',
                 'kar', 're', 'b1', 'b2', 'rr', 'c', 'hol', 'nzhk', 'ne', 'te',
                 'gi']

        for i in range(len(L1)):
            cur.execute("""INSERT INTO favourites
                        (user_id,week_day,date,time,type,food,libra, BG0)
                        VALUES (?,?,?,?,?,?,?,?)""", (session['user_id'],
                                                      week_day, date, time,
                                                      typ,
                                                      L1[i][0],
                                                      libra[i][0],
                                                      BG0))
            for elem, elem3 in zip(elem1, elem2):
                cur.execute(f"""SELECT {elem} FROM food
                             WHERE name = ?""", (L1[i][0],))
                elem = cur.fetchall()

                if elem[0][0] is None:
                    elem00 = '0'
                else:
                    elem00 = elem[0][0]

                cur.execute(f"""UPDATE favourites SET {elem3} = {elem00}
                             WHERE user_id = ? AND week_day = ?
                             AND date = ?
                             AND time = ?
                             AND type = ? AND food = ? AND libra = ?""",
                            (session['user_id'], week_day, date, time, typ,
                             L1[i][0], libra[i][0]))

        cur.execute(""" UPDATE favourites SET micr = '' """)
        cur.execute("""DELETE FROM basket WHERE user_id = ?""",
                    (session['user_id'],))
        con.commit()
        con.close()
    return redirect(url_for('news'))


@app.route('/activity')
@login_required
def activity():
    # Страница физической активности
    path = os.path.dirname(os.path.abspath(__file__))
    db_8 = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db_8)
    cur = con.cursor()
    cur.execute("""SELECT date,time,min,type,user_id
                    FROM activity WHERE user_id = ?""", (session['user_id'],))
    Act = cur.fetchall()
    cur.execute("""SELECT date,time,hour,type,user_id
                    FROM sleep WHERE user_id = ?""", (session['user_id'],))
    Sleep = cur.fetchall()
    con.close()
    return render_template('activity.html', Act=Act, Sleep=Sleep)


@app.route('/add_activity', methods=['POST'])
@login_required
def add_activity():
    # Добавляем нагрузку в базу данных
    if request.method == 'POST':

        date = datetime.datetime.strptime(request.form['calendar'], "%Y-%m-%d")
        date = date.strftime("%d.%m.%Y")

        min1 = request.form['min']
        type1 = request.form['type1']
        if type1 == '1':
            type1 = 'Ходьба'
        elif type1 == '2':
            type1 = 'Зарядка'
        elif type1 == '3':
            type1 = 'Спорт'
        elif type1 == '4':
            type1 = 'Уборка в квартире'
        elif type1 == '5':
            type1 = 'Работа в огороде'
        else:
            type1 = 'Сон'
        time1 = request.form['timer']
        path = os.path.dirname(os.path.abspath(__file__))
        db_9 = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db_9)
        cur = con.cursor()

        if type1 == 'Сон':
            cur.execute("""INSERT INTO sleep (user_id,date,time,hour,type)
                        VALUES(?,?,?,?,?)""",
                        (session['user_id'], date, time1, min1, type1))
        else:
            cur.execute("""INSERT INTO activity
                        (user_id,date,time,min,type,empty)
                        VALUES(?,?,?,?,?,?)""",
                        (session['user_id'], date, time1, min1, type1, ' '))
        con.commit()
        con.close()

    return redirect(url_for('activity'))


@app.route('/lk')
@login_required
def lk():
    # Выводим названия блюд (дневник на текущую неделю)
    session['username'] = current_user.username
    session['user_id'] = current_user.id
    session['date'] = datetime.datetime.today().date()
    
    td = datetime.datetime.today().date()
    if td.strftime("%A") == 'Monday':
        delta = datetime.timedelta(0)
    elif td.strftime("%A") == 'Tuesday':
        delta = datetime.timedelta(1)
    elif td.strftime("%A") == 'Wednesday':
        delta = datetime.timedelta(2)
    elif td.strftime("%A") == 'Thursday':
        delta = datetime.timedelta(3)
    elif td.strftime("%A") == 'Friday':
        delta = datetime.timedelta(4)
    elif td.strftime("%A") == 'Saturday':
        delta = datetime.timedelta(5)
    else:
        delta = datetime.timedelta(6)

    m = td - delta
    M = m.strftime("%d.%m.%Y")
    t = m + datetime.timedelta(1)
    T = t.strftime("%d.%m.%Y")
    w = m + datetime.timedelta(2)
    W = w.strftime("%d.%m.%Y")
    tr = m + datetime.timedelta(3)
    TR = tr.strftime("%d.%m.%Y")
    fr = m + datetime.timedelta(4)
    FR = fr.strftime("%d.%m.%Y")
    st = m + datetime.timedelta(5)
    ST = st.strftime("%d.%m.%Y")
    sd = m + datetime.timedelta(6)
    SD = sd.strftime("%d.%m.%Y")

    path = os.path.dirname(os.path.abspath(__file__))
    db_10 = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db_10)
    cur = con.cursor()

    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Понедельник',
                                  'Завтрак', M))
    MondayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""",
                (session['user_id'], 'Понедельник', 'Обед', M))
    MondayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""",
                (session['user_id'], 'Понедельник', 'Ужин', M))
    MondayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Понедельник',
                                  'Перекус', M))
    MondayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Вторник',
                                  'Завтрак', T))
    TuesdayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Вторник',
                                  'Обед', T))
    TuesdayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Вторник',
                                  'Ужин', T))
    TuesdayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Вторник',
                                  'Перекус', T))
    TuesdayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Среда',
                                  'Завтрак', W))
    WednesdayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Среда',
                                  'Обед', W))
    WednesdayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Среда',
                                  'Ужин', W))
    WednesdayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Среда',
                                  'Перекус', W))
    WednesdayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Четверг',
                                  'Завтрак', TR))
    ThursdayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Четверг',
                                  'Обед', TR))
    ThursdayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Четверг',
                                  'Ужин', TR))
    ThursdayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Четверг',
                                  'Перекус', TR))
    ThursdayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Пятница',
                                  'Завтрак', FR))
    FridayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Пятница',
                                  'Обед', FR))
    FridayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Пятница',
                                  'Ужин', FR))
    FridayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type FROM favourites
                WHERE user_id = ? AND week_day = ?
                AND type = ?
                AND date= ?""", (session['user_id'], 'Пятница',
                                 'Перекус', FR))
    FridayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Суббота',
                                  'Завтрак', ST))
    SaturdayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Суббота',
                                  'Обед', ST))
    SaturdayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Суббота',
                                  'Ужин', ST))
    SaturdayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type = ?
                AND date = ?""", (session['user_id'], 'Суббота',
                                  'Перекус', ST))
    SaturdayP = cur.fetchall()

    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type =?
                AND date = ?""", (session['user_id'], 'Воскресенье',
                                  'Завтрак', SD))
    SundayZ = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type =?
                AND date = ?""", (session['user_id'], 'Воскресенье',
                                  'Обед', SD))
    SundayO = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type =?
                AND date = ?""", (session['user_id'], 'Воскресенье',
                                  'Ужин', SD))
    SundayY = cur.fetchall()
    cur.execute(""" SELECT food,week_day,time,date,type
                FROM favourites WHERE user_id = ?
                AND week_day = ?
                AND type =?
                AND date = ?""", (session['user_id'], 'Воскресенье',
                                  'Перекус', SD))
    SundayP = cur.fetchall()
    cur.execute("""SELECT date FROM full_days
                   WHERE id = ?""", (session['user_id'],))
    date1 = cur.fetchall()
    cur.execute("""SELECT doc FROM user
                   WHERE id = ?""", (session['user_id'],))
    doc = cur.fetchall()
    con.close()

    list1 = dict()
    for i in range(len(date1)):
        list1.update({
            f'{date1[i][0]}': '',
        })

    return render_template('bootstrap_lk.html', name=session['username'],
                           MondayZ=MondayZ,
                           MondayO=MondayO,
                           MondayY=MondayY,
                           MondayP=MondayP,
                           TuesdayZ=TuesdayZ,
                           TuesdayO=TuesdayO,
                           TuesdayY=TuesdayY,
                           TuesdayP=TuesdayP,
                           WednesdayZ=WednesdayZ,
                           WednesdayO=WednesdayO,
                           WednesdayY=WednesdayY,
                           WednesdayP=WednesdayP,
                           ThursdayZ=ThursdayZ,
                           ThursdayO=ThursdayO,
                           ThursdayY=ThursdayY,
                           ThursdayP=ThursdayP,
                           FridayZ=FridayZ,
                           FridayO=FridayO,
                           FridayY=FridayY,
                           FridayP=FridayP,
                           SaturdayZ=SaturdayZ,
                           SaturdayO=SaturdayO,
                           SaturdayY=SaturdayY,
                           SaturdayP=SaturdayP,
                           SundayZ=SundayZ,
                           SundayO=SundayO,
                           SundayY=SundayY,
                           SundayP=SundayP,
                           m=M,
                           t=T,
                           w=W,
                           tr=TR,
                           fr=FR,
                           st=ST,
                           sd=SD,
                           list1=list1,
                           doc=doc)


@app.route('/delete', methods=['POST'])
@login_required
def delete():
    # Удаление данных из дневника приемов пищи за неделю
    if request.method == 'POST':
        path = os.path.dirname(os.path.abspath(__file__))
        db_11 = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db_11)
        cur = con.cursor()
        L = request.form.getlist('checked')

        for i in range(len(L)):
            L1 = L[i].split('//')

            cur.execute('''DELETE FROM favourites WHERE food = ?
                        AND date = ?
                        AND time = ?
                        AND type = ?
                        AND user_id = ?''', (L1[0], L1[1], L1[2], L1[3],
                                             session['user_id']))
            cur.execute("""INSERT INTO deleted
                        (id, date, time, type, additional)
                        VALUES (?,?,?,?,?)""", (session['user_id'],
                                                L1[1], L1[2],
                                                'Прием пищи', L1[3]))
        con.commit()
        con.close()
    return redirect(url_for('lk'))


@app.route('/remove', methods=['POST'])
@login_required
def remove():
    # Удаление данных из физической активности за неделю
    if request.method == 'POST':
        path = os.path.dirname(os.path.abspath(__file__))
        db_12 = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db_12)
        cur = con.cursor()
        L = request.form.getlist('selected')
        for i in range(len(L)):
            L1 = L[i].split('/')
            if L1[3] != 'Сон':
                cur.execute('''DELETE FROM activity WHERE date = ?
                            AND time = ?
                            AND min = ?
                            AND type = ?
                            AND user_id = ?''', (L1[0], L1[1], L1[2], L1[3],
                                                 session['user_id']))
                cur.execute("""INSERT INTO deleted
                            (id,date,time,type,additional)
                            VALUES (?,?,?,?,?)""", (session["user_id"], L1[0],
                                                    L1[1],
                                                    'Физическая активность',
                                                    L1[3]+', '+L1[2]+' минут'))
            else:
                cur.execute('''DELETE FROM sleep WHERE date = ?
                            AND time = ?
                            AND hour = ?
                            AND type = ?
                            AND user_id = ?''', (L1[0], L1[1], L1[2], L1[3],
                                                 session['user_id']))
                cur.execute("""INSERT INTO deleted
                            (id,date,time,type,additional)
                            VALUES (?,?,?,?,?)""", (session['user_id'], L1[0],
                                                    L1[1],
                                                    'Физическая активность',
                                                    L1[3]+', '+L1[2]+' часов'))
        con.commit()
        con.close()
    return redirect(url_for('activity'))


@app.route('/arch')
@login_required
def arch():
    # Архив за все время
    path = os.path.dirname(os.path.abspath(__file__))
    db_13 = os.path.join(path, 'diacompanion.db')

    con = sqlite3.connect(db_13)
    cur = con.cursor()
    cur.execute(
        """SELECT date,time,type,food,libra,carbo,prot,fat,energy
           FROM favourites WHERE user_id = ?""", (session['user_id'],))
    L = cur.fetchall()
    con.close()
    tbl = pd.DataFrame(L, columns=['Дата', 'Время', 'Прием пищи', 'Продукты',
                                   'Граммы', 'Углеводы', 'Белки', 'Жиры',
                                   'ККал'])
    tbl["Дата1"] = \
        pd.to_datetime(tbl['Дата'], format='%d.%m.%Y')
    tbl = tbl.sort_values(by="Дата1")
    tbl = tbl.drop(["Дата1"], axis=1)
    tbl = \
        tbl.groupby(
            ['Дата',
             'Время',
             'Прием пищи']).agg({
                 "Продукты": lambda tags: "br".join(tags),
                 "Граммы": lambda tags: "br".join(tags),
                 "Углеводы": lambda tags: "br".join(tags),
                 "Белки": lambda tags: "br".join(tags),
                 "Жиры": lambda tags: "br".join(tags),
                 "ККал":
                 lambda tags: "br".join(tags)}).reset_index()
    tbl = tbl.to_html(classes='table table-hover', index=False,
                      justify='left').replace('br', '</p>')
    tbl = tbl.replace('<thead>', '<thead class="thead-light">')
    tbl = tbl.replace('<table border="1" class="dataframe table table-hover">',
                      '<table class="table table-hover" aria-busy="false">')
    tbl = tbl.replace('<th>index</th>',
                      '<th>Дата</th><th>Время</th><th>Прием пищи</th>')
    tbl = tbl.replace('<th>Прием пищи</th>',
                      '<th style="white-space:nowrap;">Прием пищи</th>')
    tbl = tbl.replace('<th>ККал</th>',
                      '<th style="white-space:nowrap;">ККал</th>')
    tbl = tbl.replace('<td>', '<td class="align-middle">')

    return render_template('arch.html', tbl=tbl)


@app.route('/days', methods=['GET', 'POST'])
@login_required
def days():
    # Список полных дней
    if request.method == 'POST':
        days1 = request.form.getlist("full_days")
        day_s = days1[0].split(",")
        path = os.path.dirname(os.path.abspath(__file__))
        db_14 = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db_14)
        cur = con.cursor()
        for i in range(len(day_s)):
            cur.execute("""INSERT INTO full_days (id,date) VALUES (?,?)""",
                        (session["user_id"], day_s[i]))
        con.commit()
        con.close()
    return redirect(url_for('lk'))


def do_tb():
    global fio
    # Получили все необходимые данные из базы данных
    path = os.path.dirname(os.path.abspath(__file__))
    db_15 = os.path.join(path, 'diacompanion.db')
    model = os.path.join(path, 'model.model')
    con = sqlite3.connect(db_15)
    cur = con.cursor()
    cur.execute('''SELECT date,time,type,
                    food,libra,carbo,prot,
                    fat,energy,micr,water,mds,kr,pv,ok,
                    zola,na,k,ca,mg,p,fe,a,kar,re,b1,b2,
                    rr,c,hol,nzhk,ne,te,gi FROM favourites
                    WHERE user_id = ?''', (session['user_id'],))
    L = cur.fetchall()
    cur.execute('''SELECT date,time,min,type,empty FROM activity
                       WHERE user_id = ?''', (session['user_id'],))
    L1 = cur.fetchall()
    cur.execute('''SELECT date,time,hour FROM sleep
                        WHERE user_id =?''', (session['user_id'],))
    L2 = cur.fetchall()

    cur.execute('''SELECT DISTINCT date FROM
                        favourites WHERE user_id = ?''', (session['user_id'],))
    date = cur.fetchall()
    cur.execute("""SELECT DISTINCT date FROM full_days WHERE id = ?""",
                (session["user_id"],))
    full_days = cur.fetchall()
    cur.execute("""SELECT*FROM deleted WHERE id = ?""",
                (session["user_id"],))
    deleted = cur.fetchall()
    cur.execute("""SELECT date,time,type,BG0,gi,carbo,prot,kr FROM favourites
                       WHERE user_id = ?""", (session["user_id"],))
    tb = cur.fetchall()
    cur.execute('''SELECT username1 FROM user WHERE id = ?''',
                (session['user_id'],))
    fio = cur.fetchall()
    cur.execute('''SELECT BMI FROM user WHERE id = ?''',
                (session['user_id'],))
    BMI0 = cur.fetchall()
    BMI0 = BMI0[0][0]
    con.close()

    # Приемы пищи
    food_weight = pd.DataFrame(L, columns=['Дата', 'Время', 'Прием пищи',
                                           'Продукт', 'Масса, гр',
                                           'Углеводы, гр',
                                           'Белки, гр', 'Жиры, гр',
                                           'ККал',
                                           'Микроэлементы', 'Вода, в г',
                                           'МДС, в г',
                                           'Крахмал, в г', 'Пищ вол, в г',
                                           'Орган кислота, в г',
                                           'Зола, в г',
                                           'Натрий, в мг', 'Калий, в мг',
                                           'Кальций, в мг',
                                           'Магний, в мг', 'Фосфор, в мг',
                                           'Железо, в мг',
                                           'Ретинол, в мкг',
                                           'Каротин, в мкг',
                                           'Ретин экв, в мкг',
                                           'Тиамин, в мг',
                                           'Рибофлавин, в мг',
                                           'Ниацин, в мг',
                                           'Аскорб кисл, в мг',
                                           'Холестерин, в мг',
                                           'НЖК, в г',
                                           'Ниационвый эквивалент, в мг',
                                           'Токоферол эквивалент, в мг',
                                           'Гликемический индекс'])

    list_of = ['Масса, гр', 'Углеводы, гр',
               'Белки, гр', 'Жиры, гр',
               'ККал', 'Микроэлементы', 'Вода, в г', 'МДС, в г',
               'Крахмал, в г', 'Пищ вол, в г',
               'Орган кислота, в г', 'Зола, в г',
               'Натрий, в мг', 'Калий, в мг', 'Кальций, в мг',
               'Магний, в мг', 'Фосфор, в мг', 'Железо, в мг',
               'Ретинол, в мкг', 'Каротин, в мкг',
               'Ретин экв, в мкг', 'Тиамин, в мг',
               'Рибофлавин, в мг',
               'Ниацин, в мг', 'Аскорб кисл, в мг',
               'Холестерин, в мг',
               'НЖК, в г',
               'Ниационвый эквивалент, в мг',
               'Токоферол эквивалент, в мг',
               'Гликемический индекс']

    for name1 in list_of:
        for i in range(len(food_weight[name1])):
            food_weight[name1][i] = \
                food_weight[name1][i].replace('.', ',') + '\t'

    a = \
        food_weight.groupby(
            ['Дата',
             'Время',
             'Прием пищи']).agg({
                 "Продукт": lambda tags: '\n'.join(tags),
                 "Масса, гр": lambda tags: '\n'.join(tags),
                 "Углеводы, гр": lambda tags:
                 '\n'.join(tags),
                 "Белки, гр": lambda tags: '\n'.join(tags),
                 "Жиры, гр": lambda tags: '\n'.join(tags),
                 "ККал": lambda tags: '\n'.join(tags),
                 "Микроэлементы": lambda tags:
                 '\n'.join(tags),
                 "Вода, в г": lambda tags: '\n'.join(tags),
                 "МДС, в г": lambda tags: '\n'.join(tags),
                 "Крахмал, в г": lambda tags:
                 '\n'.join(tags),
                 "Пищ вол, в г": lambda tags:
                 '\n'.join(tags),
                 "Орган кислота, в г": lambda tags:
                 '\n'.join(tags),
                 "Зола, в г": lambda tags:
                 '\n'.join(tags),
                 "Натрий, в мг": lambda tags:
                 '\n'.join(tags),
                 "Калий, в мг": lambda tags:
                 '\n'.join(tags),
                 "Кальций, в мг": lambda tags:
                 '\n'.join(tags),
                 "Магний, в мг": lambda tags:
                 '\n'.join(tags),
                 "Фосфор, в мг": lambda tags:
                 '\n'.join(tags),
                 "Железо, в мг": lambda tags:
                 '\n'.join(tags),
                 "Ретинол, в мкг": lambda tags:
                 '\n'.join(tags),
                 "Каротин, в мкг": lambda tags:
                 '\n'.join(tags),
                 "Ретин экв, в мкг": lambda tags:
                 '\n'.join(tags),
                 "Тиамин, в мг": lambda tags:
                 '\n'.join(tags),
                 "Рибофлавин, в мг": lambda tags:
                 '\n'.join(tags),
                 "Ниацин, в мг": lambda tags:
                 '\n'.join(tags),
                 "Аскорб кисл, в мг": lambda tags:
                 '\n'.join(tags),
                 "Холестерин, в мг": lambda tags:
                 '\n'.join(tags),
                 "НЖК, в г": lambda tags:
                 '\n'.join(tags),
                 "Ниационвый эквивалент, в мг":
                 lambda tags: '\n'.join(tags),
                 "Токоферол эквивалент, в мг": lambda tags:
                 '\n'.join(tags),
                 "Гликемический индекс": lambda tags:
                 '\n'.join(tags)}).reset_index()

    a["Дата1"] = \
        pd.to_datetime(a['Дата'], format='%d.%m.%Y')
    a = a.sort_values(by="Дата1")
    a = a.drop(["Дата1"], axis=1)

    # Добавляем нумерацию блюд
    for i1 in range(len(a['Продукт'])):
        row = a['Продукт'][i1].split('\n')
        for i in range(len(row)):
            row[i] = f'{i + 1}. ' + row[i]
        row = '\n'.join(row)
        a['Продукт'][i1] = row

    if len(a['Продукт']) == 0:
        a = pd.DataFrame({'Дата': [''], 'Время': [''], 'Прием пищи': [''],
                          'Продукт': [''],
                          'Масса, гр': [''], 'Углеводы, гр': [''],
                          'Белки, гр': [''], 'Жиры, гр': [''],
                          'ККал': [''], 'Микроэлементы': [''],
                          'Вода, в г': [''], 'МДС, в г': [''],
                          'Крахмал, в г': [''], 'Пищ вол, в г': [''],
                          'Орган кислота, в г': [''],
                          'Зола, в г': [''], 'Натрий, в мг': [''],
                          'Калий, в мг': [''], 'Кальций, в мг': [''],
                          'Магний, в мг': [''], 'Фосфор, в мг': [''],
                          'Железо, в мг': [''], 'Ретинол, в мкг': [''],
                          'Каротин, в мкг': [''],
                          'Ретин экв, в мкг': [''],
                          'Тиамин, в мг': [''],
                          'Рибофлавин, в мг': [''], 'Ниацин, в мг': [''],
                          'Аскорб кисл, в мг': [''],
                          'Холестерин, в мг': [''], 'НЖК, в г': [''],
                          'Ниациновый эквивалент, в мг': [''],
                          'Токоферол эквивалент, в мг': [''],
                          'Гликемический индекс': ['']})

    # Физическая активность
    activity1 = pd.DataFrame(L1, columns=['Дата', 'Время',
                                          'Длительность, мин.',
                                          'Тип нагрузки', 'Пустое'])

    activity2 = activity1.groupby(['Дата']).agg({
        'Время': lambda tags: '\n'.join(tags),
        'Длительность, мин.': lambda tags: '\n'.join(tags),
        'Тип нагрузки': lambda tags: '\n'.join(tags),
        'Пустое': lambda tags: '\n'.join(tags)})
    # Сон
    sleep1 = pd.DataFrame(L2, columns=['Дата', 'Время',
                                       'Длительность, ч.'])

    sleep2 = \
        sleep1.groupby(
            ['Дата']).agg({'Время': lambda tags: '\n'.join(tags),
                           'Длительность, ч.': lambda tags: '\n'.join(tags)})
    luck = pd.merge(left=activity2,
                    right=sleep2,
                    on="Дата", how='outer')

    luck["Дата1"] = pd.to_datetime(luck.index, format='%d.%m.%Y')
    luck = luck.sort_values(by="Дата1")

    if len(luck.index) > 0:
        start1 = luck.index[0]
        end1 = luck.index[len(luck.index) - 1]
        start1 = datetime.datetime.strptime(start1, '%d.%m.%Y')
        end1 = datetime.datetime.strptime(end1, '%d.%m.%Y')
        start1 = start1.strftime('%m/%d/%Y')
        end1 = end1.strftime('%m/%d/%Y')

        luck = luck.drop(["Дата1"], axis=1)

        ranges = pd.date_range(start=start1, end=end1)
        ranges1 = ranges.to_pydatetime()
        new_ranges = []
        for i in range(len(ranges1)):
            new_ranges.append(ranges1[i].strftime('%d.%m.%Y'))

        luck = luck.reindex(new_ranges)
    else:
        luck = luck.drop(["Дата1"], axis=1)

    # Список полных дней
    full_days1 = pd.DataFrame(full_days, columns=['Дата'])
    full_days1['Дата1'] = pd.to_datetime(full_days1['Дата'],
                                         format='%d.%m.%Y')
    full_days1 = full_days1.sort_values(by='Дата1')
    full_days1 = full_days1.drop(['Дата1'], axis=1)

    # Список удаленных записей
    delet_ed = pd.DataFrame(deleted, columns=['id', 'Дата', 'Время', 'Тип',
                                              'Подробности'])
    delet_ed = delet_ed.drop(['id'], axis=1)
    delet_ed['Дата1'] = pd.to_datetime(delet_ed['Дата'], format='%d.%m.%Y')
    delet_ed = delet_ed.sort_values(by='Дата1')
    delet_ed = delet_ed.drop(['Дата1'], axis=1)

    # Предсказываем сахар
    tb = pd.DataFrame(tb, columns=['date', 'time', 'types_food_n', 'BG0',
                                   'GI', 'carbo', 'prot', 'kr'])

    tb['GI'] = pd.to_numeric(tb['GI'], downcast='float')
    tb['carbo'] = pd.to_numeric(tb['carbo'], downcast='float')
    tb['prot'] = pd.to_numeric(tb['prot'], downcast='float')
    tb['kr'] = pd.to_numeric(tb['kr'], downcast='float')
    tb['BG0'] = pd.to_numeric(tb['BG0'], downcast='float')

    tb = tb.groupby(['date', 'time', 'types_food_n', 'BG0'],
                    as_index=False).sum()
    tb['GL'] = tb['GI']*tb['carbo']/100
    tb['DateTime'] = tb['date'] + ' ' + tb['time']
    tb['DateTime'] = pd.to_datetime(tb['DateTime'], format='%d.%m.%Y %H:%M')
    tb = tb.drop(['date', 'time', 'GI'], axis=1)
    prot = list()
    for i in range(len(tb['DateTime'])):
        start_date = tb['DateTime'][i]
        mask = (tb['DateTime']
                <= start_date) & (tb['DateTime']
                                  >= (start_date
                                      - pd.Timedelta(value=6, unit='h')))
        prot_b6h = tb.loc[mask]['prot'].aggregate(np.sum)
        prot.append(prot_b6h)
    tb.insert(7, 'prot_b6h', prot, True)
    tb = tb.drop(['prot'], axis=1)
    BMI = list()
    for i in range(len(tb['DateTime'])):
        BMI.append(BMI0)
        if tb['types_food_n'][i] == 'Завтрак':
            tb['types_food_n'][i] = 1
        elif tb['types_food_n'][i] == 'Обед':
            tb['types_food_n'][i] = 2
        elif tb['types_food_n'][i] == 'Ужин':
            tb['types_food_n'][i] = 3
        else:
            tb['types_food_n'][i] = 4
    tb.insert(7, 'BMI', BMI, True)
    tb = tb.reindex(columns=["DateTime", "BG0", "GL", "carbo", "prot_b6h",
                             "types_food_n", "kr", "BMI"])
    predict = list()
    for i in range(len(tb['DateTime'])):
        best_model = xgb.Booster()
        best_model.load_model(model)
        core_features = ["BG0", "gl", "carbo",
                         "prot_b6h", "types_food_n", "kr", "BMI"]
        X_test = [tb.iloc[i, 1:7].values.tolist()]
        predicted = best_model.predict(xgb.DMatrix(np.array(X_test)))
        predict.append(predicted[0])
    tb.insert(3, 'Предсказанный сахар после', predict, True)

    date3 = list()
    time3 = list()
    tb['Прием пищи'] = tb['types_food_n']
    for i in range(len(tb['DateTime'])):
        date3.append(tb['DateTime'][i].strftime('%d.%m.%Y'))
        time3.append(tb['DateTime'][i].strftime('%H:%m'))
        if tb['Прием пищи'][i] == 1:
            tb['Прием пищи'][i] = 'Завтрак'
        elif tb['Прием пищи'][i] == 2:
            tb['Прием пищи'][i] = 'Обед'
        elif tb['Прием пищи'][i] == 3:
            tb['Прием пищи'][i] = 'Ужин'
        else:
            tb['Прием пищи'][i] = 'Перекус'
    tb.insert(0, 'Дата', date3, True)
    tb.insert(1, 'Время', time3, True)
    tb = tb.drop(['DateTime'], axis=1)
    tb = tb.drop(['types_food_n'], axis=1)

    tb['Сахар до'] = tb['BG0']
    tb = tb.drop(['BG0'], axis=1)
    tb['Гликемическая нагрузка'] = tb['GL']
    tb = tb.drop(['GL'], axis=1)
    tb = tb.drop(['carbo'], axis=1)
    tb = tb.drop(['prot_b6h'], axis=1)
    tb = tb.drop(['kr'], axis=1)
    tb = tb.drop(["BMI"], axis=1)
    tb = tb[['Дата', 'Время', 'Прием пищи', 'Сахар до',
             'Предсказанный сахар после', 'Гликемическая нагрузка']]

    # Создаем общий Excel файл
    dirname = os.path.dirname(__file__)
    filename = os.path.join(dirname, '%s.xlsx' % session["username"])
    writer = pd.ExcelWriter(filename,
                            engine='xlsxwriter',
                            options={'strings_to_numbers': True,
                                     'default_date_format': 'dd/mm/yy'})
    a.to_excel(writer, sheet_name='Приемы пищи', startrow=0, startcol=0)
    tb.to_excel(writer, sheet_name='Предсказание сахара',
                startrow=0, startcol=0)
    luck.to_excel(writer, sheet_name='Физическая нагрузка и сон',
                  startrow=0, startcol=1)
    full_days1.to_excel(writer, sheet_name='Список полных дней',
                        startrow=2, startcol=-1)
    delet_ed.to_excel(writer, sheet_name='Удаленные записи',
                      startrow=2, startcol=-1, header=False)
    writer.close()

    # Редактируем оформление приемов пищи
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Приемы пищи']
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True)
            cell.alignment = cell.alignment.copy(vertical='center')

    for b in ['F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
              'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
              'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']:
        for i in range(2, (len(a['Микроэлементы']) + 2)):
            k = i
            cs = sheet['%s' % b + str(k)]
            cs.alignment = cs.alignment.copy(horizontal='left')

    for c in ['B', 'C', 'D']:
        for i in range(2, (len(a['Микроэлементы']) + 2)):
            k = i
            cs = sheet['%s' % c + str(k)]
            cs.alignment = cs.alignment.copy(horizontal='center')

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 16
    sheet.column_dimensions['E'].width = 50
    sheet.column_dimensions['F'].width = 13
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 20
    sheet.column_dimensions['M'].width = 20
    sheet.column_dimensions['N'].width = 20
    sheet.column_dimensions['O'].width = 20
    sheet.column_dimensions['P'].width = 20
    sheet.column_dimensions['R'].width = 20
    sheet.column_dimensions['S'].width = 20
    sheet.column_dimensions['T'].width = 20
    sheet.column_dimensions['O'].width = 20
    sheet.column_dimensions['U'].width = 20
    sheet.column_dimensions['V'].width = 20
    sheet.column_dimensions['W'].width = 20
    sheet.column_dimensions['X'].width = 20
    sheet.column_dimensions['Y'].width = 20
    sheet.column_dimensions['Z'].width = 20
    sheet.column_dimensions['Q'].width = 20
    sheet.column_dimensions['AA'].width = 20
    sheet.column_dimensions['AB'].width = 20
    sheet.column_dimensions['AC'].width = 20
    sheet.column_dimensions['AD'].width = 20
    sheet.column_dimensions['AE'].width = 20
    sheet.column_dimensions['AF'].width = 20
    sheet.column_dimensions['AG'].width = 30
    sheet.column_dimensions['AH'].width = 30
    sheet.column_dimensions['AI'].width = 23

    b1 = ws['B1']
    b1.fill = PatternFill("solid", fgColor="fafad2")
    c1 = ws['C1']
    c1.fill = PatternFill("solid", fgColor="fafad2")
    d1 = ws['D1']
    d1.fill = PatternFill("solid", fgColor="fafad2")
    e1 = ws['E1']
    e1.fill = PatternFill("solid", fgColor="fafad2")
    f1 = ws['F1']
    f1.fill = PatternFill("solid", fgColor="fafad2")
    g1 = ws['G1']
    g1.fill = PatternFill("solid", fgColor="fafad2")
    h1 = ws['H1']
    h1.fill = PatternFill("solid", fgColor="fafad2")
    i1 = ws['I1']
    i1.fill = PatternFill("solid", fgColor="fafad2")
    j1 = ws['J1']
    j1.fill = PatternFill("solid", fgColor="fafad2")
    m1 = ws['M1']
    m1.fill = PatternFill("solid", fgColor="fafad2")
    n1 = ws['N1']
    n1.fill = PatternFill("solid", fgColor="fafad2")
    o1 = ws['O1']
    o1.fill = PatternFill("solid", fgColor="fafad2")
    p1 = ws['P1']
    p1.fill = PatternFill("solid", fgColor="fafad2")
    q1 = ws['Q1']
    q1.fill = PatternFill("solid", fgColor="fafad2")
    r1 = ws['R1']
    r1.fill = PatternFill("solid", fgColor="fafad2")
    s1 = ws['S1']
    s1.fill = PatternFill("solid", fgColor="fafad2")
    t1 = ws['T1']
    t1.fill = PatternFill("solid", fgColor="fafad2")
    u1 = ws['U1']
    u1.fill = PatternFill("solid", fgColor="fafad2")
    v1 = ws['V1']
    v1.fill = PatternFill("solid", fgColor="fafad2")
    w1 = ws['W1']
    w1.fill = PatternFill("solid", fgColor="fafad2")
    x1 = ws['X1']
    x1.fill = PatternFill("solid", fgColor="fafad2")
    y1 = ws['Y1']
    y1.fill = PatternFill("solid", fgColor="fafad2")
    z1 = ws['Z1']
    z1.fill = PatternFill("solid", fgColor="fafad2")
    aa1 = ws['AA1']
    aa1.fill = PatternFill("solid", fgColor="fafad2")
    ab1 = ws['AB1']
    ab1.fill = PatternFill("solid", fgColor="fafad2")
    ac1 = ws['AC1']
    ac1.fill = PatternFill("solid", fgColor="fafad2")
    ad1 = ws['AD1']
    ad1.fill = PatternFill("solid", fgColor="fafad2")
    ae1 = ws['AE1']
    ae1.fill = PatternFill("solid", fgColor="fafad2")
    af1 = ws['AF1']
    af1.fill = PatternFill("solid", fgColor="fafad2")
    ah1 = ws['AH1']
    ah1.fill = PatternFill("solid", fgColor="fafad2")
    ag1 = ws['AG1']
    ag1.fill = PatternFill("solid", fgColor="fafad2")
    ws['AH1'].fill = PatternFill("solid", fgColor="fafad2")
    ws['L1'].fill = PatternFill("solid", fgColor="fafad2")
    ws['AI1'].fill = PatternFill("solid", fgColor="fafad2")

    i = 1
    for num in range(1, len(a['Микроэлементы']) + 1):
        if ws[f'B{num + 1}'].value != ws[f'B{num}'].value:
            if i % 2 == 0:
                ws[f'B{num + 1}'].fill = \
                    PatternFill("solid", fgColor="f0f8ff")
                i = i + 1
            else:
                ws[f'B{num + 1}'].fill = \
                    PatternFill("solid", fgColor="f0fff0")
                i = i + 1
        else:
            ws[f'B{num + 1}']._style = ws[f'B{num}']._style

    for i in ["C", "D", "E", "F", "G", "H", "I", "J", 'L', 'M', 'N', 'O',
              'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W',
              'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF',
              'AG', 'AH', 'AI']:
        for num in range(1, len(a['Микроэлементы']) + 2):
            cell = ws[f'B{num}']
            ws[f'{i}{num}'].fill = \
                PatternFill("solid", fgColor=cell.fill.start_color.index)

    thin_border = Border(left=Side(style='hair'),
                         right=Side(style='hair'),
                         top=Side(style='hair'),
                         bottom=Side(style='hair'))

    no_border = Border(left=Side(border_style=None),
                       right=Side(border_style=None),
                       top=Side(border_style=None),
                       bottom=Side(border_style=None))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    merged_cells_range = ws.merged_cells.ranges

    for merged_cell in merged_cells_range:
        merged_cell.shift(0, 2)
    ws.insert_rows(1, 2)

    # Разделяем основные показатели и микроэлементы
    ws['K3'].value = ''
    for i in range(len(a['Микроэлементы']) + 3):
        i1 = str(i + 1)
        ws[f'K{i1}'].border = no_border

    # Убираем форматирование первого столбца A1 и последнего AI
    for i in range(len(a['Микроэлементы']) + 3):
        i1 = str(i + 1)
        ws[f'A{i1}'].border = no_border
        ws[f'A{i1}'].value = ''

    # Оформляем верхушки
    ws['A2'] = 'Приемы пищи'
    ws['A1'] = '%s' % fio[0][0]
    sheet.merge_cells('A1:AI1')
    ws['A2'].border = thin_border
    ws['A2'].fill = PatternFill("solid", fgColor="fafad2")
    ws['A2'].font = Font(bold=True)
    sheet.merge_cells('A2:AI2')

    length2 = str(len(a['Микроэлементы']) + 5)
    length3 = str(len(a['Микроэлементы']) + 6)
    sheet.merge_cells('C%s:E%s' % (length3, length3))
    ws['A%s' % length2] = 'Срденее по дням'
    ws['A%s' % length2].font = Font(bold=True)
    ws['B%s' % length3] = 'Дата'
    ws['B%s' % length3].font = Font(bold=True)
    ws['A%s' % length2].border = thin_border
    ws['A%s' % length2].fill = PatternFill("solid", fgColor="fafad2")
    ws['B%s' % length3].border = thin_border
    ws['B%s' % length3].fill = PatternFill("solid", fgColor="fafad2")
    ws['C%s' % length3].border = thin_border
    ws['C%s' % length3].fill = PatternFill("solid", fgColor="fafad2")

    # Проставляем внизу для средних по дням те же наименования,
    # что и сверху
    mean21 = ['Масса, гр', 'Углеводы, гр',
              'Белки, гр', 'Жиры, гр',
              'ККал', '', 'Вода, в г', 'МДС, в г',
              'Крахмал, в г', 'Пищ вол, в г',
              'Орган кислота, в г', 'Зола, в г',
              'Натрий, в мг', 'Калий, в мг', 'Кальций, в мг',
              'Магний, в мг', 'Фосфор, в мг', 'Железо, в мг',
              'Ретинол, в мкг', 'Каротин, в мкг',
              'Ретин экв, в мкг', 'Тиамин, в мг',
              'Рибофлавин, в мг',
              'Ниацин, в мг', 'Аскорб кисл, в мг',
              'Холестерин, в мг',
              'НЖК, в г',
              'Ниационвый эквивалент, в мг',
              'Токоферол эквивалент, в мг',
              'Гликемический индекс']
    i = 0
    for c in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
              'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
              'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']:
        ws[f'{c}%s' % length3] = mean21[i]
        ws[f'{c}%s' % length3].border = thin_border
        ws[f'{c}%s' % length3].fill = \
            PatternFill("solid", fgColor="fafad2")
        ws[f'{c}%s' % length3].font = Font(bold=True)
        i = i + 1

    # Убираем закрашенные клетки пустого столбца K
    length5 = str(len(a['Микроэлементы']) + 8 + len(date))
    ws['K%s' % length3]._style = copy(ws['K%s' % length5]._style)
    ws['K%s' % length3].border = no_border

    # Выводим скользящее среднее
    date1 = []
    for i in range(len(date)):
        date1.append(date[i][0])

    date2 = pd.DataFrame({'Дата': date1})

    date2['Дата1'] = pd.to_datetime(date2['Дата'], format='%d.%m.%Y')
    date2 = date2.sort_values(by=['Дата1'])
    date2 = date2.drop('Дата1', axis=1)
    date = date2.values.tolist()

    path = os.path.dirname(os.path.abspath(__file__))
    db_16 = os.path.join(path, 'diacompanion.db')
    con = sqlite3.connect(db_16)
    cur = con.cursor()
    i = 7
    for d in date:
        sheet['B%s' % str(len(a['Микроэлементы']) + i)] = d[0]
        cur.execute('''SELECT avg(libra), avg(carbo), avg(prot), avg(fat),
                           avg(energy), avg(water), avg(mds),
                           avg(kr), avg(pv), avg(ok), avg(zola), avg(na),
                           avg(k), avg(ca), avg(mg), avg(p), avg(fe),
                           avg(a), avg(kar), avg(re), avg(b1), avg(b2),
                           avg(rr), avg(c), avg(hol), avg(nzhk), avg(ne),
                           avg(te), avg(gi) FROM favourites
                           WHERE user_id = ?
                           AND date = ? ''', (session['user_id'], d[0]))
        avg = cur.fetchall()
        sheet['F%s' % str(len(a['Микроэлементы']) + i)] = avg[0][0]
        sheet['G%s' % str(len(a['Микроэлементы']) + i)] = avg[0][1]
        sheet['H%s' % str(len(a['Микроэлементы']) + i)] = avg[0][2]
        sheet['I%s' % str(len(a['Микроэлементы']) + i)] = avg[0][3]
        sheet['J%s' % str(len(a['Микроэлементы']) + i)] = avg[0][4]
        sheet['L%s' % str(len(a['Микроэлементы']) + i)] = avg[0][5]
        sheet['M%s' % str(len(a['Микроэлементы']) + i)] = avg[0][6]
        sheet['N%s' % str(len(a['Микроэлементы']) + i)] = avg[0][7]
        sheet['O%s' % str(len(a['Микроэлементы']) + i)] = avg[0][8]
        sheet['P%s' % str(len(a['Микроэлементы']) + i)] = avg[0][9]
        sheet['Q%s' % str(len(a['Микроэлементы']) + i)] = avg[0][10]
        sheet['R%s' % str(len(a['Микроэлементы']) + i)] = avg[0][11]
        sheet['S%s' % str(len(a['Микроэлементы']) + i)] = avg[0][12]
        sheet['T%s' % str(len(a['Микроэлементы']) + i)] = avg[0][13]
        sheet['U%s' % str(len(a['Микроэлементы']) + i)] = avg[0][14]
        sheet['V%s' % str(len(a['Микроэлементы']) + i)] = avg[0][15]
        sheet['W%s' % str(len(a['Микроэлементы']) + i)] = avg[0][16]
        sheet['X%s' % str(len(a['Микроэлементы']) + i)] = avg[0][17]
        sheet['Y%s' % str(len(a['Микроэлементы']) + i)] = avg[0][18]
        sheet['Z%s' % str(len(a['Микроэлементы']) + i)] = avg[0][19]
        sheet['AA%s' % str(len(a['Микроэлементы']) + i)] = avg[0][20]
        sheet['AB%s' % str(len(a['Микроэлементы']) + i)] = avg[0][21]
        sheet['AC%s' % str(len(a['Микроэлементы']) + i)] = avg[0][22]
        sheet['AD%s' % str(len(a['Микроэлементы']) + i)] = avg[0][23]
        sheet['AE%s' % str(len(a['Микроэлементы']) + i)] = avg[0][24]
        sheet['AF%s' % str(len(a['Микроэлементы']) + i)] = avg[0][25]
        sheet['AG%s' % str(len(a['Микроэлементы']) + i)] = avg[0][26]
        sheet['AH%s' % str(len(a['Микроэлементы']) + i)] = avg[0][27]
        sheet['AI%s' % str(len(a['Микроэлементы']) + i)] = avg[0][28]
        i = i + 1
    con.close()

    # Выравниваем скользящее среднее по левому краю
    length31 = len(a['Микроэлементы']) + 7
    length4 = len(a['Микроэлементы']) + 7 + len(date)

    for a in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
              'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
              'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']:
        for i in range(length31, length4):
            sheet[f'{a}{i}'].alignment = \
                sheet[f'{a}{i}'].alignment.copy(horizontal='left')
    for a in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
              'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
              'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF',
              'AG', 'AH', 'AI']:
        sheet[f'{a}3'].alignment = \
            sheet[f'{a}3'].alignment.copy(horizontal='left')

    ws.protection.set_password('test')
    wb.save(filename)
    wb.close()

    # Форматируем физическую активность как надо
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb['Физическая нагрузка и сон']

    for row in sheet1.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True)
            cell.alignment = cell.alignment.copy(vertical='center')
            cell.alignment = cell.alignment.copy(horizontal='left')

    for row in sheet1.iter_rows():
        for cell in row:
            cell.border = thin_border

    merged_cells_range = sheet1.merged_cells.ranges

    for merged_cell in merged_cells_range:
        merged_cell.shift(0, 2)
    sheet1.insert_rows(1, 2)

    sheet1['A1'] = '%s' % fio[0][0]
    sheet1['A2'] = 'Физическая нагрузка'
    sheet1['G2'] = 'Сон'
    sheet1.merge_cells('A1:H1')

    sheet1.column_dimensions['A'].width = 25
    sheet1.column_dimensions['B'].width = 13
    sheet1.column_dimensions['C'].width = 13
    sheet1.column_dimensions['D'].width = 20
    sheet1.column_dimensions['E'].width = 25
    sheet1.column_dimensions['F'].width = 13
    sheet1.column_dimensions['G'].width = 13
    sheet1.column_dimensions['H'].width = 20

    b1 = sheet1['B3']
    b1.fill = PatternFill("solid", fgColor="fafad2")
    c1 = sheet1['C3']
    c1.fill = PatternFill("solid", fgColor="fafad2")
    d1 = sheet1['D3']
    d1.fill = PatternFill("solid", fgColor="fafad2")
    e1 = sheet1['E3']
    e1.fill = PatternFill("solid", fgColor="fafad2")

    g1 = sheet1['G3']
    g1.fill = PatternFill("solid", fgColor="fafad2")
    sheet1['H3'].fill = PatternFill("solid", fgColor="fafad2")

    # Разделяем физическую нагрузку и сон, также убираем форматирование
    # с первого столбца A1
    # убираем мелкие дефекты
    sheet1['F3'].value = ''
    sheet1['C3'].value = 'Время'
    sheet1['G3'].value = 'Время'
    for i in range(3, len(luck['Длительность, ч.']) + 4):
        i1 = str(i)
        sheet1[f'F{i1}'].border = no_border

    for i in range(3, len(luck['Длительность, ч.']) + 4):
        i1 = str(i)
        sheet1[f'A{i1}'].border = no_border

    # Корректируем верхушки
    sheet1['A2'].fill = PatternFill("solid", fgColor="fafad2")
    sheet1['G2'].fill = PatternFill("solid", fgColor="fafad2")
    sheet1['A2'].border = thin_border
    sheet1['G2'].border = thin_border

    sheet1['A2'].font = Font(bold=True)
    sheet1['G2'].font = Font(bold=True)
    for i in range(4, len(luck['Время_x']) + 4):
        sheet1[f'B{i}'].font = Font(bold=False)

    # Закрашиваем строки через одну
    k = 1
    for i in range(4, len(luck['Длительность, ч.']) + 4):
        if k % 2 == 0:
            sheet1[f'B{i}'].fill = PatternFill('solid', fgColor='f0f8ff')
            k = k + 1
        else:
            sheet1[f'B{i}'].fill = PatternFill('solid', fgColor='f0fff0')
            k = k + 1

    for i in ["C", "D", "E", "G", "H"]:
        for num in range(4, len(luck['Длительность, ч.']) + 4):
            cell = sheet1[f'B{num}']
            sheet1[f'{i}{num}'].fill = \
                PatternFill("solid", fgColor=cell.fill.start_color.index)

    sheet1.protection.set_password('test')
    wb.save(filename)
    wb.close()

    # Форматируем список полных дней
    wb = openpyxl.load_workbook(filename)
    sheet2 = wb['Список полных дней']

    for row in sheet2.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True)
            cell.alignment = cell.alignment.copy(vertical='center')
            cell.alignment = cell.alignment.copy(horizontal='left')
            cell.border = thin_border

    sheet2['A1'] = '%s' % fio[0][0]
    sheet2['A1'].border = no_border
    sheet2['A2'] = 'Список полных дней'
    sheet2.column_dimensions['A'].width = 25
    sheet2['A2'].fill = PatternFill("solid", fgColor="fafad2")
    sheet2['A3'].fill = PatternFill("solid", fgColor="fafad2")

    sheet2['A2'].font = Font(bold=True)
    sheet2['A3'].font = Font(bold=True)

    sheet2.protection.set_password('test')
    wb.save(filename)
    wb.close()

    # Форматируем удаленные записи
    wb = openpyxl.load_workbook(filename)
    sheet3 = wb['Удаленные записи']

    sheet3["A1"] = '%s' % fio[0][0]
    sheet3["A2"].value = 'Удаленные записи'
    sheet3["A2"].font = Font(bold=True)
    sheet3["A2"].border = thin_border
    sheet3["A2"].fill = PatternFill('solid', fgColor='fafad2')

    sheet3.column_dimensions['A'].width = 25
    sheet3.column_dimensions['B'].width = 10
    sheet3.column_dimensions['C'].width = 30
    sheet3.column_dimensions['D'].width = 25

    sheet3.protection.set_password('test')
    wb.save(filename)
    wb.close()

    # Форматируем предсказание сахара
    wb = openpyxl.load_workbook(filename)
    sheet4 = wb['Предсказание сахара']

    for row in sheet4.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True)
            cell.alignment = cell.alignment.copy(vertical='center')

    for b in ["B", "C", "D", "E", "F", "G"]:
        for i in range(1, (len(tb['Прием пищи']) + 2)):
            k = i
            cs = sheet4['%s' % b + str(k)]
            cs.alignment = cs.alignment.copy(horizontal='left')

    sheet4.column_dimensions['B'].width = 15
    sheet4.column_dimensions['C'].width = 15
    sheet4.column_dimensions['D'].width = 15
    sheet4.column_dimensions['E'].width = 15
    sheet4.column_dimensions['F'].width = 30
    sheet4.column_dimensions['G'].width = 25

    sheet4['B1'].fill = PatternFill("solid", fgColor="fafad2")
    sheet4['C1'].fill = PatternFill("solid", fgColor="fafad2")
    sheet4['D1'].fill = PatternFill("solid", fgColor="fafad2")
    sheet4['E1'].fill = PatternFill("solid", fgColor="fafad2")
    sheet4['F1'].fill = PatternFill("solid", fgColor="fafad2")
    sheet4['G1'].fill = PatternFill("solid", fgColor="fafad2")

    i = 1
    for num in range(1, len(tb['Прием пищи']) + 1):
        if sheet4[f'B{num + 1}'].value != sheet4[f'B{num}'].value:
            if i % 2 == 0:
                sheet4[f'B{num + 1}'].fill = \
                    PatternFill("solid", fgColor="f0f8ff")
                i = i + 1
            else:
                sheet4[f'B{num + 1}'].fill = \
                    PatternFill("solid", fgColor="f0fff0")
                i = i + 1
        else:
            sheet4[f'B{num + 1}']._style = sheet4[f'B{num}']._style

    for i in ["C", "D", "E", "F", "G"]:
        for num in range(2, len(tb['Прием пищи']) + 2):
            cell = sheet4[f'B{num}']
            sheet4[f'{i}{num}'].fill = \
                PatternFill("solid", fgColor=cell.fill.start_color.index)

    thin_border = Border(left=Side(style='hair'),
                         right=Side(style='hair'),
                         top=Side(style='hair'),
                         bottom=Side(style='hair'))

    no_border = Border(left=Side(border_style=None),
                       right=Side(border_style=None),
                       top=Side(border_style=None),
                       bottom=Side(border_style=None))

    for row in sheet4.iter_rows():
        for cell in row:
            cell.border = thin_border

    for i in range(len(tb['Прием пищи']) + 3):
        i1 = str(i + 1)
        sheet4[f'A{i1}'].border = no_border
        sheet4[f'A{i1}'].value = ''

    merged_cells_range = sheet4.merged_cells.ranges

    for merged_cell in merged_cells_range:
        merged_cell.shift(0, 2)
    sheet4.insert_rows(1, 2)

    sheet4['A2'] = 'Предсказание сахара после приемов пищи'
    sheet4['A1'] = '%s ИМТ = %s' % (fio[0][0], BMI0)
    sheet4.merge_cells('A1:G1')
    sheet4['A2'].border = thin_border
    sheet4['A2'].fill = PatternFill("solid", fgColor="fafad2")
    sheet4['A2'].font = Font(bold=True)
    sheet4.merge_cells('A2:G2')

    sheet4.protection.set_password('test')
    wb.save(filename)
    wb.close()


def send_async_email(app, msg):
    with app.app_context():
        mail.send(msg)


@app.route('/email', methods=['GET', 'POST'])
@login_required
def email():
    # Отправляем отчет по почте отчет
    if request.method == 'POST':
        dirname = os.path.dirname(__file__)
        filename = os.path.join(dirname, '%s.xlsx' % session["username"])
        # Получили список имейлов на которые надо отправить
        mail1 = request.form.getlist('email_sendto')
        do_tb()
        # Отправляем по почте
        msg = Message('ДиаКомпаньон', sender='pochtadiacomp@gmail.com',
                      recipients=mail1)
        msg.subject = "Никнейм пользователя: %s" % session["username"]
        msg.body = 'Электронный отчет пользователя: %s' % fio[0][0]
        with app.open_resource(filename) as attach:
            msg.attach('%s.xlsx' % session["username"], 'sheet/xlsx',
                       attach.read())
        thr = Thread(target=send_async_email, args=[app, msg])
        thr.start()

        try:
            os.remove(filename)
        except FileNotFoundError:
            print('FileNotFoundError: нечего удалять')

    return redirect(url_for('lk'))


@app.route('/sendto', methods=['GET', 'POST'])
@login_required
def download_file():
    if request.method == 'POST':
        dirname = os.path.dirname(__file__)
        file_path = os.path.join(dirname, '%s.xlsx' % session["username"])
        do_tb()
    # Для Linux систем    
    # @after_this_request
    # def removing(response):
    #    os.remove(file_path)
    #    return response
    return send_file(file_path, as_attachment=True)


@app.route("/setMBI", methods=['GET', 'POST'])
@login_required
def setMBI():
    if request.method == 'POST':
        jsonBMI = request.get_json()
        path = os.path.dirname(os.path.abspath(__file__))
        db_17 = os.path.join(path, 'diacompanion.db')
        con = sqlite3.connect(db_17)
        cur = con.cursor()
        cur.execute("""UPDATE user SET BMI = ? WHERE id = ?""",
                    (jsonBMI['BMI'], session["user_id"]))
        con.commit()
        con.close()
        list2 = jsonify({"Message": "ИМТ записан"})
        response = make_response(list2, 200)
    return response


@app.route("/add_smth", methods=['GET', 'POST'])
@login_required
def add_smth():
    if request.method == 'POST':
        a101 = 'lala1'
        print(a101)
    return redirect(url_for('lk'))


if __name__ == '__main__':
    app.run(debug=True)
